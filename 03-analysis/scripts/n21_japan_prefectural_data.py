#!/usr/bin/env python3
"""
n21_japan_prefectural_data.py — 日本47都道府県パネルデータ構築 & Urban Q 計算
==============================================================================
目的:
    内閣府県民経済計算の公式 Excel ファイルから、47 都道府県の GDP、総固定資本形成 (GFCF)、
    人口のパネルデータを抽出し、都道府県レベルの Urban Q (= DeltaGDP / GFCF) を構築する。

データソース (全て内閣府 経済社会総合研究所 県民経済計算):
    - 68SNA S30 基準 (FY1955-1974): GDP, GFCF (支出側内訳)
    - 68SNA S50 基準 (FY1975-1995): GDP, GFCF (支出側内訳), 人口
    - 93SNA H21 基準 (FY1996-2009): GDP, GFCF (支出側内訳), 人口
    - 08SNA 2022 基準 (FY2011-2022): GDP, GFCF (支出側内訳), 人口
    注: FY2010 は H21 と 2022 の間の接続年で、H21 の 2009 + 2022 の 2011 から補間

入力:
    02-data/raw/japan_cab_office/*.xls, *.xlsx (内閣府ウェブサイトからダウンロード済み)

出力:
    02-data/raw/japan_prefectural_panel.csv          — 統合パネルデータ
    03-analysis/models/japan_prefectural_report.txt   — データ品質レポート + Urban Q 分析

依存: pandas, numpy, openpyxl, xlrd, scipy, statsmodels
作成日: 2026-03-22
"""

import os
import numpy as np
import pandas as pd
import xlrd
import openpyxl
from scipy import stats
import statsmodels.api as sm

# ============================================================
# パス設定
# ============================================================
PROJECT_ROOT = "/Users/andy/Desktop/Claude/urban-q-phase-transition"
CAB_OFFICE_DIR = os.path.join(PROJECT_ROOT, "02-data/raw/japan_cab_office")
RAW_OUTPUT = os.path.join(PROJECT_ROOT, "02-data/raw/japan_prefectural_panel.csv")
REPORT_OUTPUT = os.path.join(PROJECT_ROOT, "03-analysis/models/japan_prefectural_report.txt")

os.makedirs(os.path.dirname(RAW_OUTPUT), exist_ok=True)
os.makedirs(os.path.dirname(REPORT_OUTPUT), exist_ok=True)

# 報告用バッファ
report_lines = []
def rpt(s=''):
    report_lines.append(s)
    print(s)

# ============================================================
# 都道府県コード → 名称マッピング
# ============================================================
PREF_CODES = {
    '01': '北海道', '02': '青森県', '03': '岩手県', '04': '宮城県', '05': '秋田県',
    '06': '山形県', '07': '福島県', '08': '茨城県', '09': '栃木県', '10': '群馬県',
    '11': '埼玉県', '12': '千葉県', '13': '東京都', '14': '神奈川県', '15': '新潟県',
    '16': '富山県', '17': '石川県', '18': '福井県', '19': '山梨県', '20': '長野県',
    '21': '岐阜県', '22': '静岡県', '23': '愛知県', '24': '三重県', '25': '滋賀県',
    '26': '京都府', '27': '大阪府', '28': '兵庫県', '29': '奈良県', '30': '和歌山県',
    '31': '鳥取県', '32': '島根県', '33': '岡山県', '34': '広島県', '35': '山口県',
    '36': '徳島県', '37': '香川県', '38': '愛媛県', '39': '高知県', '40': '福岡県',
    '41': '佐賀県', '42': '長崎県', '43': '熊本県', '44': '大分県', '45': '宮崎県',
    '46': '鹿児島県', '47': '沖縄県',
}

PREF_EN = {
    '01': 'Hokkaido', '02': 'Aomori', '03': 'Iwate', '04': 'Miyagi', '05': 'Akita',
    '06': 'Yamagata', '07': 'Fukushima', '08': 'Ibaraki', '09': 'Tochigi', '10': 'Gunma',
    '11': 'Saitama', '12': 'Chiba', '13': 'Tokyo', '14': 'Kanagawa', '15': 'Niigata',
    '16': 'Toyama', '17': 'Ishikawa', '18': 'Fukui', '19': 'Yamanashi', '20': 'Nagano',
    '21': 'Gifu', '22': 'Shizuoka', '23': 'Aichi', '24': 'Mie', '25': 'Shiga',
    '26': 'Kyoto', '27': 'Osaka', '28': 'Hyogo', '29': 'Nara', '30': 'Wakayama',
    '31': 'Tottori', '32': 'Shimane', '33': 'Okayama', '34': 'Hiroshima', '35': 'Yamaguchi',
    '36': 'Tokushima', '37': 'Kagawa', '38': 'Ehime', '39': 'Kochi', '40': 'Fukuoka',
    '41': 'Saga', '42': 'Nagasaki', '43': 'Kumamoto', '44': 'Oita', '45': 'Miyazaki',
    '46': 'Kagoshima', '47': 'Okinawa',
}

REGION_MAP = {
    '01': '北海道', '02': '東北', '03': '東北', '04': '東北', '05': '東北',
    '06': '東北', '07': '東北', '08': '関東', '09': '関東', '10': '関東',
    '11': '関東', '12': '関東', '13': '関東', '14': '関東', '15': '中部',
    '16': '中部', '17': '中部', '18': '中部', '19': '中部', '20': '中部',
    '21': '中部', '22': '中部', '23': '中部', '24': '中部', '25': '近畿',
    '26': '近畿', '27': '近畿', '28': '近畿', '29': '近畿', '30': '近畿',
    '31': '中国', '32': '中国', '33': '中国', '34': '中国', '35': '中国',
    '36': '四国', '37': '四国', '38': '四国', '39': '四国', '40': '九州',
    '41': '九州', '42': '九州', '43': '九州', '44': '九州', '45': '九州',
    '46': '九州', '47': '九州',
}


# ============================================================
# ヘルパー関数
# ============================================================

def extract_soukatu_xls(filepath, data_row_start=6, year_row=4, code_col=0, name_col=1, data_start_col=3):
    """
    内閣府の総括表 (.xls) からデータを抽出する汎用関数。
    形式: 行=都道府県 (47行), 列=年度
    返り値: dict of {(pref_code, year): value}
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # 年度を取得
    years = []
    for j in range(data_start_col, ws.ncols):
        val = ws.cell_value(year_row, j)
        if isinstance(val, float) and val > 1900:
            years.append((j, int(val)))

    # データ抽出
    result = {}
    for i in range(data_row_start, ws.nrows):
        code = str(ws.cell_value(i, code_col)).strip()
        if code in PREF_CODES:
            # ゼロパディング
            code = code.zfill(2)
            for j, yr in years:
                val = ws.cell_value(i, j)
                if isinstance(val, (int, float)) and val != 0:
                    result[(code, yr)] = val

    return result


def extract_soukatu_xlsx(filepath, data_row_start=7, year_row=5, code_col=0, name_col=1, data_start_col=3):
    """
    内閣府の総括表 (.xlsx) からデータを抽出する汎用関数。
    形式: 行=都道府県 (47行), 列=年度
    返り値: dict of {(pref_code, year): value}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))

    # 年度を取得 (year_row は 1-indexed in the data)
    years = []
    year_data = rows[year_row - 1]  # 0-indexed
    for j, val in enumerate(year_data):
        if isinstance(val, (int, float)) and val is not None and val > 1900:
            years.append((j, int(val)))

    # データ抽出
    result = {}
    for row in rows[data_row_start - 1:]:  # 0-indexed
        if row[code_col] is None:
            continue
        code = str(row[code_col]).strip()
        if code in PREF_CODES:
            code = code.zfill(2)
            for j, yr in years:
                val = row[j] if j < len(row) else None
                if isinstance(val, (int, float)) and val is not None:
                    result[(code, yr)] = val

    return result


def extract_expenditure_s30(filepath):
    """
    S30基準の支出側データ (長形式) から総固定資本形成 (GFCF) を抽出。
    形式: 行=都道府県×項目, 列=年度
    項目番号 '4' = '(1)総固定資本形成'
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name('データ')

    # 年度を取得
    years = []
    for j in range(4, ws.ncols):
        val = ws.cell_value(3, j)
        if isinstance(val, str) and '年度' in val:
            # 昭和30年度 -> 1955
            import re
            m = re.search(r'(\d+)', val)
            if m:
                era_year = int(m.group(1))
                western_year = era_year + 1925  # 昭和
                years.append((j, western_year))
        elif isinstance(val, (int, float)) and val > 1900:
            years.append((j, int(val)))

    # 年度行から直接取得
    if not years:
        for j in range(4, ws.ncols):
            val = ws.cell_value(4, j) if ws.nrows > 4 else None
            if isinstance(val, (int, float)) and val > 1900:
                years.append((j, int(val)))

    result = {}
    for i in range(4, ws.nrows):
        code = str(ws.cell_value(i, 0)).strip().zfill(2)
        item_num = str(ws.cell_value(i, 2)).strip()
        item_name = str(ws.cell_value(i, 3)).strip()

        if code in PREF_CODES and item_num == '4':
            # (1)総固定資本形成
            for j, yr in years:
                val = ws.cell_value(i, j)
                if isinstance(val, (int, float)) and val != 0:
                    result[(code, yr)] = val

    return result


def extract_expenditure_s50_sheet(filepath):
    """
    S50基準の支出側データ (シート=年度) から総固定資本形成を抽出。
    各シートは1年度分、列8 = 総固定資本形成
    シート名: S50, 51, ..., 63, H1, 2, ..., 11
      S50=昭和50年(1975), 51=昭和51年(1976), ..., 63=昭和63年(1988)
      H1=平成1年(1989), 2=平成2年(1990), ..., 11=平成11年(1999)
    """
    wb = xlrd.open_workbook(filepath)
    result = {}

    # シート名→年度マッピング (順序に依存)
    sheet_names = wb.sheet_names()
    in_heisei = False
    for sn in sheet_names:
        year = None
        if sn.upper().startswith('S'):
            # S50 -> 昭和50年 -> 1975
            try:
                era_year = int(sn.upper().replace('S', '').strip())
                year = era_year + 1925
            except ValueError:
                pass
        elif sn.upper().startswith('H'):
            # H1 -> 平成1年 -> 1989
            in_heisei = True
            try:
                era_year = int(sn.upper().replace('H', '').strip())
                year = era_year + 1988
            except ValueError:
                pass
        else:
            try:
                num = int(sn.strip())
                if in_heisei:
                    # H1 の後は平成 (2, 3, ..., 11 = 1990, 1991, ..., 1999)
                    year = num + 1988
                else:
                    # S50 の後は昭和 (51, 52, ..., 63 = 1976, ..., 1988)
                    year = num + 1925
            except ValueError:
                pass

        if year is None:
            continue

        ws = wb.sheet_by_name(sn)
        # col 8 = 総固定資本形成 (confirmed from structure analysis)
        for i in range(8, ws.nrows):
            code = str(ws.cell_value(i, 0)).strip().zfill(2)
            if code in PREF_CODES:
                val = ws.cell_value(i, 8)
                if isinstance(val, (int, float)) and val != 0:
                    result[(code, year)] = val

    return result


def extract_expenditure_h21_sheet(filepath):
    """
    H21基準 syuyo4 (支出側詳細) から総固定資本形成を抽出。
    各シートは1年度分。col 8 = 総固定資本形成。
    """
    wb = xlrd.open_workbook(filepath)
    result = {}

    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        # シート名から年度を推定
        year = None
        import re
        m = re.search(r'(\d{4})', sn)
        if m:
            year = int(m.group(1))
        if year is None:
            continue

        for i in range(8, ws.nrows):
            code_val = ws.cell_value(i, 0)
            code = str(code_val).strip().zfill(2) if code_val not in ('', None) else ''
            if code in PREF_CODES:
                val = ws.cell_value(i, 8)
                if isinstance(val, (int, float)) and val != 0:
                    result[(code, year)] = val

    return result


def extract_expenditure_2022_sheet(filepath):
    """
    2022基準 syuyo4 (.xlsx, 支出側詳細) から総固定資本形成と支出側GDP を抽出。
    各シートは1年度分。
    col 27 = 県内総資本形成
    col 28 = 総固定資本形成 (GFCF)
    col 29 = 民間住宅
    col 30 = 民間企業設備
    col 47 = 県内総生産(支出側)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    gfcf_result = {}
    gfcf_private_housing = {}
    gfcf_private_equip = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # 年度推定
        import re
        m = re.search(r'(\d{4})', sn)
        if not m:
            continue
        year = int(m.group(1))

        for row in rows[8:]:  # 0-indexed, data starts at row 9
            if row[0] is None:
                continue
            code = str(row[0]).strip().zfill(2)
            if code in PREF_CODES:
                # col 28 = 総固定資本形成
                val = row[28] if len(row) > 28 else None
                if isinstance(val, (int, float)):
                    gfcf_result[(code, year)] = val
                # col 29 = 民間住宅
                val_h = row[29] if len(row) > 29 else None
                if isinstance(val_h, (int, float)):
                    gfcf_private_housing[(code, year)] = val_h
                # col 30 = 民間企業設備
                val_e = row[30] if len(row) > 30 else None
                if isinstance(val_e, (int, float)):
                    gfcf_private_equip[(code, year)] = val_e

    return gfcf_result, gfcf_private_housing, gfcf_private_equip


# ============================================================
# ステップ 1: GDP 抽出
# ============================================================
rpt('=' * 72)
rpt('日本 47 都道府県パネルデータ構築')
rpt('Japan 47 Prefectures Panel Data Construction')
rpt('=' * 72)
rpt()

rpt('--- ステップ 1: 県内総生産 (名目 GDP, 百万円) 抽出 ---')

gdp_data = {}

# S30 基準 (FY1955-1974)
f = os.path.join(CAB_OFFICE_DIR, 's30_30soukatu1.xls')
if os.path.exists(f):
    d = extract_soukatu_xls(f, data_row_start=6, year_row=4)
    gdp_data.update(d)
    years_s30 = sorted(set(yr for _, yr in d.keys()))
    rpt(f'  S30 基準: {len(d)} obs, FY{min(years_s30)}-{max(years_s30)}')

# S50 基準 (FY1975-1995)
f = os.path.join(CAB_OFFICE_DIR, 's50_50soukatu1.xls')
if os.path.exists(f):
    d = extract_soukatu_xls(f, data_row_start=6, year_row=4)
    # S50 とS30 の重複年は S50 を優先
    for k, v in d.items():
        gdp_data[k] = v
    years_s50 = sorted(set(yr for _, yr in d.keys()))
    rpt(f'  S50 基準: {len(d)} obs, FY{min(years_s50)}-{max(years_s50)}')

# H21 基準 (FY1996-2009) — 93SNA
f = os.path.join(CAB_OFFICE_DIR, 'h21_soukatu1_1.xls')
if os.path.exists(f):
    d = extract_soukatu_xls(f, data_row_start=6, year_row=4)
    for k, v in d.items():
        gdp_data[k] = v
    years_h21 = sorted(set(yr for _, yr in d.keys()))
    rpt(f'  H21 基準: {len(d)} obs, FY{min(years_h21)}-{max(years_h21)}')

# 2022 基準 (FY2011-2022) — 08SNA
f = os.path.join(CAB_OFFICE_DIR, 'soukatu1.xlsx')
if os.path.exists(f):
    d = extract_soukatu_xlsx(f, data_row_start=7, year_row=5)
    for k, v in d.items():
        gdp_data[k] = v
    years_2022 = sorted(set(yr for _, yr in d.keys()))
    rpt(f'  2022 基準: {len(d)} obs, FY{min(years_2022)}-{max(years_2022)}')

all_gdp_years = sorted(set(yr for _, yr in gdp_data.keys()))
rpt(f'  GDP 合計: {len(gdp_data)} obs, FY{min(all_gdp_years)}-{max(all_gdp_years)}')
rpt()

# ============================================================
# ステップ 2: 総固定資本形成 (GFCF, 百万円) 抽出
# ============================================================
rpt('--- ステップ 2: 総固定資本形成 (GFCF, 百万円) 抽出 ---')

gfcf_data = {}

# S30 基準 (FY1955-1974) — 長形式
f = os.path.join(CAB_OFFICE_DIR, 's30_30sisyutu1_s30-s49.xls')
if os.path.exists(f):
    d = extract_expenditure_s30(f)
    gfcf_data.update(d)
    if d:
        years_gfcf_s30 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  S30 基準: {len(d)} obs, FY{min(years_gfcf_s30)}-{max(years_gfcf_s30)}')
    else:
        rpt('  S30 基準: 抽出失敗')

# S50 基準 (FY1975-1995) — シート形式
f = os.path.join(CAB_OFFICE_DIR, 's50_50sisyutu1_s50-h11.xls')
if os.path.exists(f):
    d = extract_expenditure_s50_sheet(f)
    for k, v in d.items():
        gfcf_data[k] = v
    if d:
        years_gfcf_s50 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  S50 基準: {len(d)} obs, FY{min(years_gfcf_s50)}-{max(years_gfcf_s50)}')
    else:
        rpt('  S50 基準: 抽出失敗')

# H21 基準 (FY1996-2009) — syuyo4 シート形式
f = os.path.join(CAB_OFFICE_DIR, 'h21_syuyo4_1.xls')
if os.path.exists(f):
    d = extract_expenditure_h21_sheet(f)
    for k, v in d.items():
        gfcf_data[k] = v
    if d:
        years_gfcf_h21 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  H21 基準: {len(d)} obs, FY{min(years_gfcf_h21)}-{max(years_gfcf_h21)}')
    else:
        rpt('  H21 基準: 抽出失敗')

# 2022 基準 (FY2011-2022) — syuyo4.xlsx シート形式
f = os.path.join(CAB_OFFICE_DIR, 'syuyo4.xlsx')
gfcf_priv_housing = {}
gfcf_priv_equip = {}
if os.path.exists(f):
    d, dh, de = extract_expenditure_2022_sheet(f)
    for k, v in d.items():
        gfcf_data[k] = v
    gfcf_priv_housing.update(dh)
    gfcf_priv_equip.update(de)
    if d:
        years_gfcf_2022 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  2022 基準: {len(d)} obs, FY{min(years_gfcf_2022)}-{max(years_gfcf_2022)}')
    else:
        rpt('  2022 基準: 抽出失敗')

if gfcf_data:
    all_gfcf_years = sorted(set(yr for _, yr in gfcf_data.keys()))
    rpt(f'  GFCF 合計: {len(gfcf_data)} obs, FY{min(all_gfcf_years)}-{max(all_gfcf_years)}')
else:
    rpt('  警告: GFCF データ抽出失敗')
rpt()

# ============================================================
# ステップ 3: 人口 (人) 抽出
# ============================================================
rpt('--- ステップ 3: 総人口 (人) 抽出 ---')

pop_data = {}

# S50 基準 (FY1975-1995)
f = os.path.join(CAB_OFFICE_DIR, 's50_50soukatu9.xls')
if os.path.exists(f):
    d = extract_soukatu_xls(f, data_row_start=6, year_row=4)
    pop_data.update(d)
    if d:
        years_pop_s50 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  S50 基準: {len(d)} obs, FY{min(years_pop_s50)}-{max(years_pop_s50)}')

# H21 基準 (FY1996-2009)
f = os.path.join(CAB_OFFICE_DIR, 'h21_soukatu11_pop.xls')
if os.path.exists(f):
    d = extract_soukatu_xls(f, data_row_start=6, year_row=4)
    for k, v in d.items():
        pop_data[k] = v
    if d:
        years_pop_h21 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  H21 基準: {len(d)} obs, FY{min(years_pop_h21)}-{max(years_pop_h21)}')

# 2022 基準 (FY2011-2022)
f = os.path.join(CAB_OFFICE_DIR, 'soukatu9.xlsx')
if os.path.exists(f):
    d = extract_soukatu_xlsx(f, data_row_start=7, year_row=5)
    for k, v in d.items():
        pop_data[k] = v
    if d:
        years_pop_2022 = sorted(set(yr for _, yr in d.keys()))
        rpt(f'  2022 基準: {len(d)} obs, FY{min(years_pop_2022)}-{max(years_pop_2022)}')

if pop_data:
    all_pop_years = sorted(set(yr for _, yr in pop_data.keys()))
    rpt(f'  人口 合計: {len(pop_data)} obs, FY{min(all_pop_years)}-{max(all_pop_years)}')
rpt()

# ============================================================
# ステップ 4: パネルデータ統合
# ============================================================
rpt('--- ステップ 4: パネルデータ統合 ---')

# 全ての (pref_code, year) の組み合わせを収集
all_keys = set(gdp_data.keys()) | set(gfcf_data.keys()) | set(pop_data.keys())
all_prefs = sorted(set(k[0] for k in all_keys))
all_years = sorted(set(k[1] for k in all_keys))

rows = []
for code in all_prefs:
    if code not in PREF_CODES:
        continue
    for yr in all_years:
        row = {
            'pref_code': code,
            'prefecture_jp': PREF_CODES[code],
            'prefecture_en': PREF_EN.get(code, ''),
            'region': REGION_MAP.get(code, ''),
            'year': yr,
            'gdp_nominal_myen': gdp_data.get((code, yr), np.nan),
            'gfcf_nominal_myen': gfcf_data.get((code, yr), np.nan),
            'population': pop_data.get((code, yr), np.nan),
            'gfcf_private_housing_myen': gfcf_priv_housing.get((code, yr), np.nan),
            'gfcf_private_equip_myen': gfcf_priv_equip.get((code, yr), np.nan),
        }
        # GDP があるか GFCF があるか 人口がある行のみ保持
        if not (np.isnan(row['gdp_nominal_myen']) and np.isnan(row['gfcf_nominal_myen']) and np.isnan(row['population'])):
            rows.append(row)

panel = pd.DataFrame(rows)
panel = panel.sort_values(['pref_code', 'year']).reset_index(drop=True)

rpt(f'  パネル: {panel.shape[0]} rows x {panel.shape[1]} cols')
rpt(f'  都道府県: {panel.pref_code.nunique()} prefectures')
rpt(f'  年度: {panel.year.min()} - {panel.year.max()}')
rpt(f'  GDP: {panel.gdp_nominal_myen.notna().sum()} non-null')
rpt(f'  GFCF: {panel.gfcf_nominal_myen.notna().sum()} non-null')
rpt(f'  人口: {panel.population.notna().sum()} non-null')
rpt()

# ============================================================
# ステップ 5: 派生変数の計算
# ============================================================
rpt('--- ステップ 5: 派生変数の計算 ---')

# 人均 GDP (百万円)
panel['gdp_per_capita_myen'] = panel['gdp_nominal_myen'] / panel['population']

# GDP 成長率 (名目)
panel = panel.sort_values(['pref_code', 'year'])
panel['gdp_growth_nominal'] = panel.groupby('pref_code')['gdp_nominal_myen'].pct_change(fill_method=None)

# GDP 変化量 (DeltaGDP)
panel['delta_gdp_myen'] = panel.groupby('pref_code')['gdp_nominal_myen'].diff()

# GFCF / GDP 比率 (投資率)
panel['gfcf_gdp_ratio'] = panel['gfcf_nominal_myen'] / panel['gdp_nominal_myen']

# 人口成長率
panel['pop_growth'] = panel.groupby('pref_code')['population'].pct_change(fill_method=None)

# MUQ = DeltaGDP / GFCF (Marginal Urban Q)
# 解釈: 追加投資1円あたりの GDP 増加分
panel['muq'] = panel['delta_gdp_myen'] / panel['gfcf_nominal_myen']

# 3年移動平均 MUQ (ノイズ低減)
panel['muq_ma3'] = panel.groupby('pref_code')['muq'].transform(
    lambda x: x.rolling(3, center=True, min_periods=2).mean()
)

# ln 変数
panel['ln_gdp'] = np.log(panel['gdp_nominal_myen'].clip(lower=1))
panel['ln_pop'] = np.log(panel['population'].clip(lower=1))
panel['ln_gdp_pc'] = np.log(panel['gdp_per_capita_myen'].clip(lower=1e-6))

rpt(f'  MUQ (non-null): {panel.muq.notna().sum()}')
rpt(f'  MUQ mean: {panel.muq.mean():.4f}')
rpt(f'  MUQ median: {panel.muq.median():.4f}')
rpt()

# ============================================================
# ステップ 6: FY2010 補間 (H21 と 2022 のギャップ)
# ============================================================
rpt('--- ステップ 6: FY2010 補間 ---')

# FY2010 のデータが欠損している場合、FY2009 と FY2011 から線形補間
count_interpolated = 0
for code in PREF_CODES.keys():
    mask_2010 = (panel['pref_code'] == code) & (panel['year'] == 2010)
    if mask_2010.sum() == 0:
        # FY2010 行がない場合は追加
        row_2009 = panel[(panel['pref_code'] == code) & (panel['year'] == 2009)]
        row_2011 = panel[(panel['pref_code'] == code) & (panel['year'] == 2011)]
        if len(row_2009) > 0 and len(row_2011) > 0:
            new_row = {
                'pref_code': code,
                'prefecture_jp': PREF_CODES[code],
                'prefecture_en': PREF_EN.get(code, ''),
                'region': REGION_MAP.get(code, ''),
                'year': 2010,
            }
            for col in ['gdp_nominal_myen', 'gfcf_nominal_myen', 'population']:
                v2009 = row_2009[col].values[0]
                v2011 = row_2011[col].values[0]
                if np.isfinite(v2009) and np.isfinite(v2011):
                    new_row[col] = (v2009 + v2011) / 2
                else:
                    new_row[col] = np.nan
            panel = pd.concat([panel, pd.DataFrame([new_row])], ignore_index=True)
            count_interpolated += 1

if count_interpolated > 0:
    rpt(f'  FY2010 補間: {count_interpolated} prefectures')
    # 再計算
    panel = panel.sort_values(['pref_code', 'year']).reset_index(drop=True)
    panel['gdp_per_capita_myen'] = panel['gdp_nominal_myen'] / panel['population']
    panel['gdp_growth_nominal'] = panel.groupby('pref_code')['gdp_nominal_myen'].pct_change(fill_method=None)
    panel['delta_gdp_myen'] = panel.groupby('pref_code')['gdp_nominal_myen'].diff()
    panel['gfcf_gdp_ratio'] = panel['gfcf_nominal_myen'] / panel['gdp_nominal_myen']
    panel['pop_growth'] = panel.groupby('pref_code')['population'].pct_change(fill_method=None)
    panel['muq'] = panel['delta_gdp_myen'] / panel['gfcf_nominal_myen']
    panel['muq_ma3'] = panel.groupby('pref_code')['muq'].transform(
        lambda x: x.rolling(3, center=True, min_periods=2).mean()
    )
    panel['ln_gdp'] = np.log(panel['gdp_nominal_myen'].clip(lower=1))
    panel['ln_pop'] = np.log(panel['population'].clip(lower=1))
    panel['ln_gdp_pc'] = np.log(panel['gdp_per_capita_myen'].clip(lower=1e-6))
else:
    rpt('  FY2010 補間: 不要')
rpt()

# ============================================================
# ステップ 7: SNA 基準マーカー付与
# ============================================================
def assign_sna_basis(year):
    if year <= 1974:
        return '68SNA_S30'
    elif year <= 1995:
        return '68SNA_S50'
    elif year <= 2009:
        return '93SNA_H21'
    elif year == 2010:
        return 'interpolated'
    else:
        return '08SNA_2022'

panel['sna_basis'] = panel['year'].apply(assign_sna_basis)

# ============================================================
# ステップ 8: 出力
# ============================================================
rpt('--- ステップ 8: CSV 出力 ---')
panel.to_csv(RAW_OUTPUT, index=False, encoding='utf-8-sig')
rpt(f'  出力: {RAW_OUTPUT}')
rpt(f'  サイズ: {panel.shape[0]} rows x {panel.shape[1]} cols')
rpt()

# ============================================================
# ステップ 9: データ品質レポート
# ============================================================
rpt('=' * 72)
rpt('データ品質レポート')
rpt('=' * 72)
rpt()

# カバレッジ
rpt('--- カバレッジ ---')
for var, label in [('gdp_nominal_myen', 'GDP'), ('gfcf_nominal_myen', 'GFCF'), ('population', '人口')]:
    non_null = panel[var].notna()
    years_covered = sorted(panel.loc[non_null, 'year'].unique())
    if years_covered:
        rpt(f'  {label}: FY{min(years_covered)}-{max(years_covered)}, {non_null.sum()} obs ({non_null.mean()*100:.1f}%)')
    else:
        rpt(f'  {label}: データなし')

rpt()

# SNA 基準別カバレッジ
rpt('--- SNA 基準別 ---')
for basis in ['68SNA_S30', '68SNA_S50', '93SNA_H21', 'interpolated', '08SNA_2022']:
    sub = panel[panel['sna_basis'] == basis]
    if len(sub) > 0:
        rpt(f'  {basis}: {len(sub)} obs, GDP={sub.gdp_nominal_myen.notna().sum()}, '
            f'GFCF={sub.gfcf_nominal_myen.notna().sum()}, Pop={sub.population.notna().sum()}')

rpt()

# 基本統計量 (直近 2022 年クロスセクション)
rpt('--- FY2022 クロスセクション基本統計量 ---')
cs = panel[panel['year'] == 2022].copy()
if len(cs) == 47:
    rpt(f'  N = {len(cs)} prefectures')
    for var, label, unit in [
        ('gdp_nominal_myen', 'GDP (名目)', '百万円'),
        ('gfcf_nominal_myen', 'GFCF', '百万円'),
        ('population', '人口', '人'),
        ('gdp_per_capita_myen', '人均GDP', '百万円/人'),
        ('gfcf_gdp_ratio', 'GFCF/GDP', ''),
    ]:
        if cs[var].notna().sum() > 0:
            rpt(f'  {label} ({unit}):')
            rpt(f'    mean = {cs[var].mean():,.0f}, sd = {cs[var].std():,.0f}')
            rpt(f'    min = {cs[var].min():,.0f} ({cs.loc[cs[var].idxmin(), "prefecture_en"]})')
            rpt(f'    max = {cs[var].max():,.0f} ({cs.loc[cs[var].idxmax(), "prefecture_en"]})')
else:
    rpt(f'  FY2022: {len(cs)} prefectures (expected 47)')
rpt()

# ============================================================
# ステップ 10: Urban Q 分析
# ============================================================
rpt('=' * 72)
rpt('Urban Q (MUQ) 分析')
rpt('=' * 72)
rpt()

# 全国集計 MUQ 時系列
rpt('--- 全国集計 MUQ 時系列 ---')
national = panel.groupby('year').agg(
    gdp_total=('gdp_nominal_myen', 'sum'),
    gfcf_total=('gfcf_nominal_myen', 'sum'),
    pop_total=('population', 'sum'),
).dropna()

national['delta_gdp'] = national['gdp_total'].diff()
national['muq_national'] = national['delta_gdp'] / national['gfcf_total']
national['gfcf_gdp_ratio'] = national['gfcf_total'] / national['gdp_total']

# 主要期間の MUQ
periods = {
    '高度成長期 (1960-1973)': (1960, 1973),
    '安定成長期 (1974-1985)': (1974, 1985),
    'バブル期 (1986-1991)': (1986, 1991),
    '失われた10年 (1992-2002)': (1992, 2002),
    '小泉改革期 (2003-2008)': (2003, 2008),
    'アベノミクス (2013-2019)': (2013, 2019),
    'コロナ後 (2020-2022)': (2020, 2022),
}

for label, (y1, y2) in periods.items():
    sub = national[(national.index >= y1) & (national.index <= y2)]
    muq_vals = sub['muq_national'].dropna()
    muq_vals = muq_vals[np.isfinite(muq_vals)]  # inf を除外
    if len(muq_vals) > 0:
        rpt(f'  {label}: MUQ = {muq_vals.mean():.4f} (n={len(muq_vals)})')
    else:
        rpt(f'  {label}: データ不足 (SNA 基準変更の境界に注意)')
rpt()

# 都道府県別 MUQ ランキング (直近 5 年平均)
rpt('--- 都道府県 MUQ ランキング (FY2018-2022 平均) ---')
recent = panel[(panel['year'] >= 2018) & (panel['year'] <= 2022)]
pref_muq = recent.groupby(['pref_code', 'prefecture_en'])['muq'].mean().reset_index()
pref_muq = pref_muq.sort_values('muq', ascending=False)

rpt('  Top 10:')
for _, row in pref_muq.head(10).iterrows():
    rpt(f'    {row.prefecture_en:15s}: MUQ = {row.muq:+.4f}')
rpt('  Bottom 10:')
for _, row in pref_muq.tail(10).iterrows():
    rpt(f'    {row.prefecture_en:15s}: MUQ = {row.muq:+.4f}')
rpt()

# バブル期 vs 現在の GFCF/GDP 比率
rpt('--- GFCF/GDP 比率の推移 ---')
for period_label, (y1, y2) in [('高度成長 1965-1973', (1965, 1973)),
                                ('バブル期 1986-1991', (1986, 1991)),
                                ('2018-2022', (2018, 2022))]:
    sub = national[(national.index >= y1) & (national.index <= y2)]
    ratio = sub['gfcf_gdp_ratio'].dropna()
    if len(ratio) > 0:
        rpt(f'  {period_label}: GFCF/GDP = {ratio.mean():.3f}')
rpt()

# 東京 vs 地方の MUQ 比較
rpt('--- 東京 vs 非東京 MUQ 比較 ---')
for period_label, (y1, y2) in [('FY1996-2005', (1996, 2005)), ('FY2011-2022', (2011, 2022))]:
    sub = panel[(panel['year'] >= y1) & (panel['year'] <= y2)]
    tokyo = sub[sub['pref_code'] == '13']['muq'].dropna()
    non_tokyo = sub[sub['pref_code'] != '13']['muq'].dropna()
    if len(tokyo) > 0 and len(non_tokyo) > 0:
        t_stat, p_val = stats.ttest_ind(tokyo, non_tokyo, equal_var=False)
        rpt(f'  {period_label}:')
        rpt(f'    東京: MUQ = {tokyo.mean():.4f} (n={len(tokyo)})')
        rpt(f'    非東京: MUQ = {non_tokyo.mean():.4f} (n={len(non_tokyo)})')
        rpt(f'    差の検定: t = {t_stat:.3f}, p = {p_val:.4f}')
rpt()

# 人口減少と MUQ の関係
rpt('--- 人口減少と MUQ の関係 (2011-2022 パネル) ---')
sub = panel[(panel['year'] >= 2012) & (panel['year'] <= 2022)].dropna(subset=['muq', 'pop_growth'])
if len(sub) > 30:
    r, p = stats.pearsonr(sub['pop_growth'], sub['muq'])
    rpt(f'  Pearson r = {r:.4f}, p = {p:.4f}, n = {len(sub)}')

    # パネル固定効果回帰
    sub_fe = sub.copy()
    sub_fe['pop_growth_pct'] = sub_fe['pop_growth'] * 100
    # 県ダミー + 年ダミー
    year_dummies = pd.get_dummies(sub_fe['year'], prefix='yr', drop_first=True, dtype=float)
    pref_dummies = pd.get_dummies(sub_fe['pref_code'], prefix='pref', drop_first=True, dtype=float)
    X = pd.concat([sub_fe[['pop_growth_pct']], year_dummies, pref_dummies], axis=1)
    X = sm.add_constant(X)
    y = sub_fe['muq']
    valid = X.notna().all(axis=1) & y.notna()
    if valid.sum() > X.shape[1] + 5:
        model = sm.OLS(y[valid], X[valid]).fit(cov_type='cluster',
                                                cov_kwds={'groups': sub_fe.loc[valid, 'pref_code']})
        rpt(f'  固定効果回帰 (MUQ ~ pop_growth + FE):')
        rpt(f'    pop_growth 係数 = {model.params["pop_growth_pct"]:.4f}')
        rpt(f'    SE (cluster) = {model.bse["pop_growth_pct"]:.4f}')
        rpt(f'    p = {model.pvalues["pop_growth_pct"]:.4f}')
        rpt(f'    R2 = {model.rsquared:.4f}')
rpt()

# ============================================================
# ステップ 11: SNA 基準接続の注意
# ============================================================
rpt('=' * 72)
rpt('注意: SNA 基準の接続について')
rpt('=' * 72)
rpt()
rpt('本データセットは 4 つの異なる SNA 基準をチェーンリンクしている:')
rpt('  1. 68SNA S30 基準 (FY1955-1974): 旧体系、名目値')
rpt('  2. 68SNA S50 基準 (FY1975-1995): 旧体系、名目値')
rpt('  3. 93SNA H21 基準 (FY1996-2009): 新体系、連鎖方式実質あり')
rpt('  4. 08SNA 2022 基準 (FY2011-2022): 最新体系、研究開発資本化含む')
rpt()
rpt('基準間の接続は直接接合 (各基準の最新推計値を優先)。')
rpt('基準変更により以下の不連続性が存在しうる:')
rpt('  - 1974/1975 境界: 68SNA S30 → S50 (軽微)')
rpt('  - 1995/1996 境界: 68SNA → 93SNA (中程度)')
rpt('  - 2009/2010/2011 境界: 93SNA → 08SNA (中程度、R&D 資本化の影響)')
rpt('分析には期間ダミーまたは構造断点検定を推奨。')
rpt()

# レポートを書き出し
with open(REPORT_OUTPUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(report_lines))
rpt(f'レポート出力: {REPORT_OUTPUT}')

# サマリー
rpt()
rpt('=' * 72)
rpt('完了サマリー')
rpt('=' * 72)
rpt(f'パネルデータ: {RAW_OUTPUT}')
rpt(f'  {panel.shape[0]} observations')
rpt(f'  {panel.pref_code.nunique()} prefectures')
rpt(f'  FY{panel.year.min()}-{panel.year.max()}')
rpt(f'  変数: {", ".join(panel.columns[:12])} ...')
rpt(f'レポート: {REPORT_OUTPUT}')
